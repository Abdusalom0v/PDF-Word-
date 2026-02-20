"""Excel export functionality."""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class ExcelExporter:
    """Exports selected words to Excel (.xlsx) format."""

    @staticmethod
    def export(words: List[str], output_path: Path) -> bool:
        """Export words to Excel file.

        Args:
            words: List of words to export.
            output_path: Target file path for the Excel file.

        Returns:
            True if export succeeded, False otherwise.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Words"

            # Header
            ws["A1"] = "Selected Words"
            ws["A1"].font = Font(bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            for row, word in enumerate(words, start=2):
                ws.cell(row=row, column=1, value=word)

            ws.column_dimensions["A"].width = 30
            wb.save(output_path)
            return True
        except Exception:
            return False
